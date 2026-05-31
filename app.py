# v16 - car minimizing algorithm
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
import sqlite3, json, os, io, math
from datetime import datetime, date, timedelta

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', os.urandom(24).hex())
PORT = int(os.environ.get('PORT', 8080))
_base = os.path.dirname(os.path.abspath(__file__))
# Use /data volume on Railway (persistent storage), fallback to local
_data_dir = '/data' if os.path.isdir('/data') else _base
DB = os.path.join(_data_dir, 'logistics.db')

LOGISTICS_COLS = [
    {'name': 'שם',              'type': 'text',   'options': '', 'fixed': True},
    {'name': 'שם משפחה',        'type': 'text',   'options': '', 'fixed': True},
    {'name': 'צוות',            'type': 'select', 'options': '', 'fixed': True, 'teams_col': True},
    {'name': 'תאריך הגעה',      'type': 'date',   'options': '', 'fixed': True},
    {'name': 'תאריך חזרה',      'type': 'date',   'options': '', 'fixed': True},
    {'name': 'אילוצים/הערות',   'type': 'text',   'options': '', 'fixed': True},
    {'name': 'רכב חברה',        'type': 'status', 'options': 'כן,לא', 'fixed': True},
    {'name': 'רכב הלוך',        'type': 'text',   'options': '', 'fixed': True, 'auto_col': True},
    {'name': 'רכב חזור',        'type': 'text',   'options': '', 'fixed': True, 'auto_col': True},
    {'name': 'מספר חדר',        'type': 'text',   'options': '', 'fixed': True},
]
ADMIN_ONLY_COLS = set()  # ת"ז הוסר
AUTO_COLS       = {'רכב הלוך', 'רכב חזור'}

# ── DB ─────────────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as db:
        db.executescript('''
            CREATE TABLE IF NOT EXISTS lists (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                name         TEXT    NOT NULL,
                icon         TEXT    DEFAULT '🗂️',
                columns      TEXT    NOT NULL,
                teams        TEXT    DEFAULT '[]',
                car_count    INTEGER DEFAULT 4,
                car_capacity INTEGER DEFAULT 3,
                trial_start  TEXT    DEFAULT NULL,
                trial_end    TEXT    DEFAULT NULL,
                created_at   TEXT    DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS list_rows (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                list_id      INTEGER NOT NULL,
                data         TEXT    NOT NULL,
                confirmed_by TEXT    DEFAULT NULL,
                confirmed_at TEXT    DEFAULT NULL,
                created_at   TEXT    DEFAULT (datetime('now')),
                updated_at   TEXT    DEFAULT (datetime('now')),
                FOREIGN KEY (list_id) REFERENCES lists(id)
            );
            CREATE TABLE IF NOT EXISTS changes_log (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                row_id       INTEGER,
                list_id      INTEGER,
                action       TEXT,
                field_name   TEXT,
                old_value    TEXT,
                new_value    TEXT,
                done_by      TEXT,
                created_at   TEXT    DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS approved_gaps (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                list_id      INTEGER,
                gap_key      TEXT,
                approved_by  TEXT,
                approved_at  TEXT,
                UNIQUE(list_id, gap_key)
            );
            CREATE TABLE IF NOT EXISTS car_drivers (
                id        INTEGER PRIMARY KEY AUTOINCREMENT,
                list_id   INTEGER,
                direction TEXT,
                car_name  TEXT,
                driver    TEXT,
                UNIQUE(list_id, direction, car_name)
            );
            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT
            );
            CREATE TABLE IF NOT EXISTS users (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                username   TEXT NOT NULL UNIQUE,
                password   TEXT NOT NULL,
                role       TEXT DEFAULT 'user',
                approved   INTEGER DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now'))
            );
        ''')
        import os as _os
        defaults = [
            ('password',       _os.environ.get('USER_PASSWORD',  'logistics2026')),
            ('admin_password', _os.environ.get('ADMIN_PASSWORD', 'admin2026')),
        ]
        for k, v in defaults:
            if not db.execute("SELECT 1 FROM app_settings WHERE key=?", (k,)).fetchone():
                db.execute("INSERT INTO app_settings VALUES (?,?)", (k, v))
        db.commit()

init_db()

# ── Helpers ────────────────────────────────────────────────────────────────────
def get_setting(k):
    with get_db() as db:
        r = db.execute("SELECT value FROM app_settings WHERE key=?", (k,)).fetchone()
    return r['value'] if r else None

def is_admin():
    return session.get('role') == 'admin'

def normalize_header(h):
    if h is None: return ''
    s = str(h).strip()
    s = s.replace(chr(92), '/')
    while '  ' in s:
        s = s.replace('  ', ' ')
    return s

def cell_to_str(val):
    if val is None: return ''
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    if isinstance(val, date):     return val.strftime('%Y-%m-%d')
    if isinstance(val, bool):     return 'כן' if val else 'לא'
    s = str(val).strip()
    if s.endswith('.0') and s[:-2].lstrip('-').isdigit(): return s[:-2]
    return s

def log_change(db, row_id, list_id, action, field=None, old=None, new=None, by=None):
    db.execute(
        "INSERT INTO changes_log (row_id,list_id,action,field_name,old_value,new_value,done_by) VALUES (?,?,?,?,?,?,?)",
        (row_id, list_id, action, field, old, new, by or session.get('username','?'))
    )

def mask_admin(data):
    return {k: ('***' if k in ADMIN_ONLY_COLS else v) for k,v in data.items()}

def get_list(lid):
    with get_db() as db:
        r = db.execute("SELECT * FROM lists WHERE id=?", (lid,)).fetchone()
    return dict(r) if r else None

def get_rows(lid, admin=False):
    with get_db() as db:
        rows = db.execute("SELECT * FROM list_rows WHERE list_id=? ORDER BY id", (lid,)).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        raw = json.loads(d['data'])
        d['data'] = raw if admin else mask_admin(raw)
        result.append(d)
    return result

# ── Auth ───────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        pw = request.form.get('password','').strip()
        name = request.form.get('name','').strip()
        adm = get_setting('admin_password') or 'admin2026'
        reg = get_setting('password') or 'logistics2026'
        if pw == adm:
            session.update({'auth':True,'role':'admin','username':name or 'אדמין'})
            return redirect(url_for('index'))
        elif pw == reg:
            session.update({'auth':True,'role':'user','username':name or 'משתמש'})
            return redirect(url_for('index'))
        else:
            error = 'סיסמה שגויה'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def auth_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**kw):
        if not session.get('auth'): return redirect(url_for('login'))
        return f(*a,**kw)
    return d

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def d(*a,**kw):
        if not session.get('auth') or not is_admin():
            return jsonify({'error':'אין הרשאה'}),403
        return f(*a,**kw)
    return d

# ── Pages ──────────────────────────────────────────────────────────────────────
@app.route('/')
@auth_required
def index():
    from flask import make_response
    with get_db() as db:
        lists = db.execute("SELECT * FROM lists ORDER BY created_at DESC").fetchall()
    resp = make_response(render_template('index.html', lists=lists,
                           username=session.get('username',''), is_admin=is_admin()))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    return resp

@app.route('/list/<int:lid>')
@auth_required
def list_view(lid):
    lst = get_list(lid)
    if not lst: return redirect(url_for('index'))
    lst['columns']      = json.loads(lst['columns'])
    lst['teams']        = json.loads(lst['teams'])
    lst['teams_json']   = json.dumps(lst['teams'],  ensure_ascii=False)
    lst['trial_start']  = lst['trial_start'] or ''
    lst['trial_end']    = lst['trial_end']   or ''
    return render_template('list.html', lst=lst,
                           username=session.get('username',''), is_admin=is_admin())

# ── API: Lists ─────────────────────────────────────────────────────────────────
@app.route('/api/lists', methods=['GET'])
@auth_required
def api_get_lists():
    with get_db() as db:
        rows = db.execute("SELECT * FROM lists ORDER BY created_at DESC").fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/lists', methods=['POST'])
@admin_required
def api_create_list():
    d = request.json
    cols = json.dumps(LOGISTICS_COLS, ensure_ascii=False)
    with get_db() as db:
        cur = db.execute(
            "INSERT INTO lists (name,icon,columns,teams,car_count,car_capacity) VALUES (?,?,?,?,?,?)",
            (d['name'], d.get('icon','🗂️'), cols, '[]', 4, 3)
        )
        db.commit()
    return jsonify({'id': cur.lastrowid})

@app.route('/api/lists/<int:lid>', methods=['PUT'])
@admin_required
def api_update_list(lid):
    d = request.json
    with get_db() as db:
        db.execute("UPDATE lists SET name=?,icon=? WHERE id=?", (d['name'],d.get('icon','🗂️'),lid))
        db.commit()
    return jsonify({'ok':True})

@app.route('/api/lists/<int:lid>', methods=['DELETE'])
@admin_required
def api_delete_list(lid):
    with get_db() as db:
        db.execute("DELETE FROM list_rows WHERE list_id=?",(lid,))
        db.execute("DELETE FROM changes_log WHERE list_id=?",(lid,))
        db.execute("DELETE FROM approved_gaps WHERE list_id=?",(lid,))
        db.execute("DELETE FROM car_drivers WHERE list_id=?",(lid,))
        db.execute("DELETE FROM lists WHERE id=?",(lid,))
        db.commit()
    return jsonify({'ok':True})

@app.route('/api/lists/<int:lid>/config', methods=['GET'])
@auth_required
def api_get_config(lid):
    lst = get_list(lid)
    if not lst: return jsonify({'error':'not found'}),404
    return jsonify({'teams':json.loads(lst['teams']),
                    'car_count':lst['car_count'],'car_capacity':lst['car_capacity'],
                    'trial_start':lst['trial_start'] or '','trial_end':lst['trial_end'] or ''})

@app.route('/api/lists/<int:lid>/config', methods=['POST'])
@admin_required
def api_save_config(lid):
    d = request.json
    new_teams = d.get('teams',[])
    with get_db() as db:
        old = db.execute("SELECT teams FROM lists WHERE id=?", (lid,)).fetchone()
        old_teams = set(json.loads(old['teams'])) if old else set()
        added_teams = [t for t in new_teams if t not in old_teams]
        db.execute("""UPDATE lists SET teams=?,car_count=?,car_capacity=?,trial_start=?,trial_end=?
                      WHERE id=?""",
                   (json.dumps(new_teams, ensure_ascii=False),
                    int(d.get('car_count',4)), int(d.get('car_capacity',3)),
                    d.get('trial_start','') or None, d.get('trial_end','') or None, lid))
        # Auto-create placeholder rows for new teams
        if added_teams:
            cols = LOGISTICS_COLS
            for team in added_teams:
                data = {c['name']:'' for c in cols}
                data['צוות'] = team
                cur = db.execute("INSERT INTO list_rows (list_id,data) VALUES (?,?)",
                                 (lid, json.dumps(data, ensure_ascii=False)))
                log_change(db, cur.lastrowid, lid, f'שורה אוטומטית לצוות {team}', by='מערכת')
        db.commit()
    return jsonify({'ok':True,'added_teams':added_teams})

# ── API: Rows ──────────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/rows', methods=['GET'])
@auth_required
def api_get_rows(lid):
    return jsonify(get_rows(lid, is_admin()))

@app.route('/api/lists/<int:lid>/rows', methods=['POST'])
@auth_required
def api_create_row(lid):
    data = request.json
    with get_db() as db:
        cur = db.execute("INSERT INTO list_rows (list_id,data) VALUES (?,?)",
                         (lid, json.dumps(data, ensure_ascii=False)))
        log_change(db, cur.lastrowid, lid, 'נוצרה שורה')
        db.commit()
    return jsonify({'id':cur.lastrowid})

@app.route('/api/rows/<int:rid>', methods=['PUT'])
@auth_required
def api_update_row(rid):
    new_data = request.json
    now = datetime.now().isoformat()
    with get_db() as db:
        old = db.execute("SELECT * FROM list_rows WHERE id=?", (rid,)).fetchone()
        if old:
            old_data = json.loads(old['data'])
            if not is_admin():
                for c in ADMIN_ONLY_COLS:
                    if c in old_data: new_data[c] = old_data[c]
            for k in new_data:
                if str(old_data.get(k,'')) != str(new_data.get(k,'')):
                    log_change(db, rid, old['list_id'], 'עדכון', k,
                               str(old_data.get(k,'')), str(new_data.get(k,'')))
        db.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                   (json.dumps(new_data, ensure_ascii=False), now, rid))
        db.commit()
    return jsonify({'ok':True})

@app.route('/api/rows/<int:rid>/confirm', methods=['POST'])
@auth_required
def api_confirm_row(rid):
    now = datetime.now().isoformat()
    username = session.get('username','?')
    with get_db() as db:
        row = db.execute("SELECT * FROM list_rows WHERE id=?", (rid,)).fetchone()
        if not row: return jsonify({'ok':False}),404
        if row['confirmed_by']:
            db.execute("UPDATE list_rows SET confirmed_by=NULL,confirmed_at=NULL WHERE id=?", (rid,))
            log_change(db, rid, row['list_id'], 'בוטל אישור')
            db.commit()
            return jsonify({'confirmed':False})
        db.execute("UPDATE list_rows SET confirmed_by=?,confirmed_at=? WHERE id=?", (username,now,rid))
        log_change(db, rid, row['list_id'], 'אושר')
        db.commit()
    return jsonify({'confirmed':True,'by':username,'at':now})

@app.route('/api/rows/<int:rid>', methods=['DELETE'])
@auth_required
def api_delete_row(rid):
    with get_db() as db:
        row = db.execute("SELECT list_id FROM list_rows WHERE id=?", (rid,)).fetchone()
        if row: log_change(db, rid, row['list_id'], 'נמחקה שורה')
        db.execute("DELETE FROM list_rows WHERE id=?", (rid,))
        db.commit()
    return jsonify({'ok':True})

@app.route('/api/lists/<int:lid>/rows/bulk-delete', methods=['POST'])
@auth_required
def api_bulk_delete(lid):
    ids = request.json.get('ids',[])
    deleted = 0
    with get_db() as db:
        for rid in ids:
            row = db.execute("SELECT list_id FROM list_rows WHERE id=? AND list_id=?", (rid,lid)).fetchone()
            if row:
                log_change(db, rid, lid, 'נמחקה שורה (מרובה)')
                db.execute("DELETE FROM list_rows WHERE id=?", (rid,))
                deleted += 1
        db.commit()
    return jsonify({'ok':True,'deleted':deleted})

# ── API: Validate ──────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/validate', methods=['GET'])
@auth_required
def api_validate(lid):
    lst = get_list(lid)
    if not lst: return jsonify({'issues':[]})
    teams        = json.loads(lst['teams'])
    car_count    = lst['car_count']
    car_capacity = lst['car_capacity']
    total_seats  = car_count * car_capacity
    trial_start  = lst['trial_start'] or ''
    trial_end    = lst['trial_end']   or ''

    with get_db() as db:
        raw_rows = db.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
        approved = {r['gap_key'] for r in db.execute(
            "SELECT gap_key FROM approved_gaps WHERE list_id=?", (lid,)).fetchall()}

    people = []
    for r in raw_rows:
        d = json.loads(r['data'])
        people.append({
            'id':       r['id'],
            'name':     (d.get('שם','') + ' ' + d.get('שם משפחה','')).strip(),
            'team':     d.get('צוות','').strip(),
            'arrival':  d.get('תאריך הגעה','').strip(),
            'departure':d.get('תאריך חזרה','').strip(),
            'has_car':  str(d.get('רכב חברה','')).lower() in ['כן','yes','true','1'],
        })

    issues = []
    seen_keys = set()  # dedup

    def add_issue(key, **kwargs):
        if key not in seen_keys and key not in approved:
            seen_keys.add(key)
            issues.append({'key': key, **kwargs})

    if trial_start and trial_end:
        try:
            t_start = date.fromisoformat(trial_start)
            t_end   = date.fromisoformat(trial_end)
        except:
            t_start = t_end = None

        if t_start and t_end:
            all_dates = []
            d = t_start
            while d <= t_end:
                all_dates.append(d); d += timedelta(days=1)

            for day in all_dates:
                ds = day.isoformat()
                present = []
                for p in people:
                    try:
                        arr = date.fromisoformat(p['arrival'])  if p['arrival']  else None
                        dep = date.fromisoformat(p['departure']) if p['departure'] else None
                        if arr and dep and arr <= day <= dep:
                            present.append(p)
                    except: pass

                # Team coverage
                if teams:
                    for team in teams:
                        if not any(p['team'] == team for p in present):
                            # Find candidates: same team, present on adjacent days
                            candidates = [p['name'] for p in people
                                          if p['team'] == team and p['name']]
                            add_issue(f'cov_{ds}_{team}',
                                type='no_coverage', severity='error',
                                date=ds, team=team,
                                message=f'אין נציג מצוות "{team}" בתאריך {ds}',
                                approvable=True,
                                fix_type='add_person',
                                fix_data={'team': team, 'date': ds,
                                          'candidates': list(set(candidates))[:8]})

                # Car capacity per day
                arriving   = [p for p in people if p['arrival']   == ds and not p['has_car']]
                departing  = [p for p in people if p['departure']  == ds and not p['has_car']]

                if len(arriving) > total_seats:
                    extra = math.ceil((len(arriving) - total_seats) / car_capacity)
                    add_issue(f'car_arr_{ds}',
                        type='car_overflow', severity='error',
                        date=ds, direction='הגעה',
                        message=f'הגעה {ds}: {len(arriving)} נוסעים, {total_seats} מקומות. חסרים {extra} רכבים',
                        approvable=False,
                        fix_type='add_cars',
                        fix_data={'extra_cars': extra, 'direction': 'arrive'})

                if len(departing) > total_seats:
                    extra = math.ceil((len(departing) - total_seats) / car_capacity)
                    add_issue(f'car_ret_{ds}',
                        type='car_overflow', severity='error',
                        date=ds, direction='חזרה',
                        message=f'חזרה {ds}: {len(departing)} נוסעים, {total_seats} מקומות. חסרים {extra} רכבים',
                        approvable=False,
                        fix_type='add_cars',
                        fix_data={'extra_cars': extra, 'direction': 'return'})

            # Coverage gaps: someone leaves, no one from their team stays
            for p in people:
                if not p['departure'] or not p['team']: continue
                try: dep_date = date.fromisoformat(p['departure'])
                except: continue
                if dep_date >= t_end: continue
                next_day = dep_date + timedelta(days=1)
                team_next = []
                for other in people:
                    if other['id'] == p['id'] or other['team'] != p['team']: continue
                    try:
                        arr2 = date.fromisoformat(other['arrival'])  if other['arrival']  else None
                        dep2 = date.fromisoformat(other['departure']) if other['departure'] else None
                        if arr2 and dep2 and arr2 <= next_day <= dep2:
                            team_next.append(other)
                    except: pass
                if not team_next:
                    candidates = [o['name'] for o in people
                                  if o['team'] == p['team'] and o['id'] != p['id'] and o['name']]
                    add_issue(f'gap_{dep_date}_{p["team"]}',
                        type='coverage_gap', severity='warning',
                        date=dep_date.isoformat(), team=p['team'], person=p['name'],
                        message=f'{p["name"]} (צוות {p["team"]}) עוזב ב-{dep_date} ואין מחליף',
                        approvable=True,
                        fix_type='extend_or_add',
                        fix_data={'team': p['team'], 'date': dep_date.isoformat(),
                                  'next_day': next_day.isoformat(),
                                  'candidates': list(set(candidates))[:8],
                                  'person_id': p['id']})

    # Car balance: same count of unique cars must go and return
    no_car_people = [p for p in people if not p['has_car']]
    arrive_count  = len({p['name'] for p in no_car_people if p['arrival']})
    return_count  = len({p['name'] for p in no_car_people if p['departure']})
    if arrive_count and return_count:
        cars_needed_arrive = math.ceil(arrive_count  / car_capacity)
        cars_needed_return = math.ceil(return_count  / car_capacity)
        if cars_needed_arrive != cars_needed_return:
            add_issue('car_balance',
                type='car_balance', severity='warning',
                message=f'מאזן רכבים: {cars_needed_arrive} רכבים נדרשים להגעה אך {cars_needed_return} לחזרה — מספר הרכבים חייב להיות שווה',
                approvable=False,
                fix_type='car_balance_info',
                fix_data={'arrive': cars_needed_arrive, 'return': cars_needed_return,
                          'needed': max(cars_needed_arrive, cars_needed_return)})

    row_issues = {}
    for iss in issues:
        if iss.get('fix_data',{}).get('person_id'):
            pid = iss['fix_data']['person_id']
            row_issues.setdefault(str(pid),[]).append(iss['message'])

    return jsonify({'issues': issues, 'row_issues': row_issues})

# ── API: Approve Gap ───────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/gaps/approve', methods=['POST'])
@admin_required
def api_approve_gap(lid):
    key = request.json.get('key')
    with get_db() as db:
        db.execute("INSERT OR REPLACE INTO approved_gaps (list_id,gap_key,approved_by,approved_at) VALUES (?,?,?,?)",
                   (lid, key, session.get('username'), datetime.now().isoformat()))
        db.commit()
    return jsonify({'ok':True})

# ── API: Fix Issue ─────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/fix', methods=['POST'])
@admin_required
def api_fix_issue(lid):
    d = request.json
    fix_type = d.get('fix_type')
    fix_data = d.get('fix_data', {})

    if fix_type == 'add_cars':
        extra = int(fix_data.get('extra_cars', 1))
        with get_db() as db:
            lst = db.execute("SELECT car_count FROM lists WHERE id=?", (lid,)).fetchone()
            new_count = (lst['car_count'] or 4) + extra
            db.execute("UPDATE lists SET car_count=? WHERE id=?", (new_count, lid))
            db.commit()
        return jsonify({'ok':True, 'message':f'עודכן ל-{new_count} רכבים', 'new_car_count': new_count})

    if fix_type == 'add_person':
        team      = fix_data.get('team','')
        arr_date  = fix_data.get('date','')
        full_name = fix_data.get('full_name','').strip()
        if not full_name:
            return jsonify({'ok':False, 'needs_input':True, 'prompt': f'שם מלא של האדם מצוות {team} להוסיף לתאריך {arr_date}:'})
        parts = full_name.split(' ', 1)
        data = {c['name']:'' for c in LOGISTICS_COLS}
        data['שם']         = parts[0]
        data['שם משפחה']   = parts[1] if len(parts) > 1 else ''
        data['צוות']        = team
        data['תאריך הגעה']  = arr_date
        data['תאריך חזרה']  = arr_date
        with get_db() as db:
            cur = db.execute("INSERT INTO list_rows (list_id,data) VALUES (?,?)",
                             (lid, json.dumps(data, ensure_ascii=False)))
            log_change(db, cur.lastrowid, lid, f'תיקון אוטומטי: הוסף {full_name} לצוות {team}', by=session.get('username'))
            db.commit()
        return jsonify({'ok':True, 'message':f'{full_name} נוסף/ה לצוות {team} בתאריך {arr_date}', 'row_id': cur.lastrowid})

    if fix_type == 'extend_or_add':
        action    = fix_data.get('action')    # 'extend' or 'add'
        person_id = fix_data.get('person_id')
        next_day  = fix_data.get('next_day','')
        team      = fix_data.get('team','')
        full_name = fix_data.get('full_name','').strip()

        if action == 'add':
            if not full_name:
                return jsonify({'ok':False,'needs_input':True,
                                'prompt':f'שם מלא של האדם מצוות {team} להוסיף מ-{next_day}:'})
            parts = full_name.split(' ', 1)
            data = {c['name']:'' for c in LOGISTICS_COLS}
            data['שם']='  '.join(parts[:1]); data['שם משפחה']=parts[1] if len(parts)>1 else ''
            data['צוות']=team; data['תאריך הגעה']=next_day; data['תאריך חזרה']=next_day
            with get_db() as db:
                cur = db.execute("INSERT INTO list_rows (list_id,data) VALUES (?,?)",
                                 (lid, json.dumps(data, ensure_ascii=False)))
                log_change(db, cur.lastrowid, lid, f'תיקון: הוסף {full_name} מ-{next_day}', by=session.get('username'))
                db.commit()
            return jsonify({'ok':True,'message':f'{full_name} נוסף/ה מ-{next_day}'})

    return jsonify({'ok':False,'error':'סוג תיקון לא מוכר'})

# ── helpers ────────────────────────────────────────────────────────────────────
import re as _re

def _norm_date(v):
    v = str(v).strip()
    if not v or v.lower() in ['none','null','']: return ''
    m = _re.match(r'^(\d{1,2})[/\.](\d{1,2})(?:[/\.](\d{2,4}))?$', v)
    if m:
        d,mo,yr = m.group(1),m.group(2),m.group(3) or ''
        return f"{d.zfill(2)}/{mo.zfill(2)}" + (f"/{yr}" if yr else '')
    m = _re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', v)
    if m:
        return f"{m.group(3).zfill(2)}/{m.group(2).zfill(2)}/{m.group(1)}"
    m = _re.match(r'^(\d{1,2})-(\d{1,2})-(\d{2,4})$', v)
    if m:
        return f"{m.group(1).zfill(2)}/{m.group(2).zfill(2)}/{m.group(3)}"
    return v

def _gf(d, *keys):
    for k in keys:
        v = str(d.get(k,'')).strip()
        if v and v.lower() not in ['none','null','']: return v
    return ''

def _get_field(d, *keys): return _gf(d, *keys)

def _get_people(db, lid):
    rows = db.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
    people = []
    for row in rows:
        d = json.loads(row['data'])
        if _gf(d,'רכב חברה').lower() in ['כן','yes','true','1']: continue
        full = (_gf(d,'שם')+' '+_gf(d,'שם משפחה')).strip()
        if not full: continue
        people.append({
            'full':      full,
            'id':        row['id'],
            'arrival':   _norm_date(_gf(d,'תאריך הגעה','הגעה')),
            'departure': _norm_date(_gf(d,'תאריך חזרה','חזרה')),
            'car_a':     _gf(d,'רכב הלוך'),
            'car_r':     _gf(d,'רכב חזור'),
        })
    return people

def _car_people(lid):
    conn = get_db()
    try:
        return _get_people(conn, lid)
    finally:
        conn.close()

# ── API: Debug dates ──────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/debug-dates', methods=['GET'])
@admin_required
def api_debug_dates(lid):
    try:
        people = _car_people(lid)
        from collections import defaultdict
        groups = defaultdict(list)
        for p in people:
            k = f"{p['arrival'] or '?'}→{p['departure'] or '?'}"
            groups[k].append(p['full'])
        return jsonify({'ok':True,'count':len(people),
                        'groups':[{'key':k,'people':v} for k,v in sorted(groups.items())]})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)})

# ── API: Step 1 — שיבוץ אחיד ──────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/assign-step1', methods=['POST'])
@admin_required
def api_assign_step1(lid):
    from collections import defaultdict
    try:
        lst = get_list(lid)
        if not lst: return jsonify({'ok':False,'error':'רשימה לא נמצאה'})
        cap       = max(1, lst['car_capacity'])
        car_count = max(1, lst['car_count'])
        now       = datetime.now().isoformat()

        conn = get_db()
        try:
            # 1. Reset all assignments
            rows = conn.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
            for row in rows:
                d = json.loads(row['data'])
                if _gf(d,'רכב חברה').lower() in ['כן','yes','true','1']: continue
                if d.get('רכב הלוך') or d.get('רכב חזור'):
                    d['רכב הלוך'] = ''; d['רכב חזור'] = ''
                    conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                 (json.dumps(d,ensure_ascii=False), now, row['id']))
            conn.execute("DELETE FROM car_drivers WHERE list_id=?", (lid,))
            conn.commit()

            # 2. Get fresh people data
            people = _get_people(conn, lid)

            # 3. Group by (arrive_date, return_date)
            combo = defaultdict(list)
            for p in people:
                if p['arrival'] and p['departure']:
                    combo[(p['arrival'], p['departure'])].append(p)

            # 4. Assign to pool cars
            avail = [f'רכב {i}' for i in range(1, car_count+1)]
            car_idx = 0; assigned = 0; groups_info = []

            for (arr,ret), grp in sorted(combo.items(), key=lambda x: -len(x[1])):
                groups_info.append({'arr':arr,'ret':ret,'count':len(grp),
                                    'names':[p['full'] for p in grp]})
                pool = list(grp)
                while pool:
                    if car_idx >= len(avail): break
                    cname = avail[car_idx]; car_idx += 1
                    drv = pool.pop(0)
                    # Write driver (both directions)
                    r = conn.execute("SELECT data FROM list_rows WHERE id=?", (drv['id'],)).fetchone()
                    d2 = json.loads(r['data']); d2['רכב הלוך']=cname; d2['רכב חזור']=cname
                    conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                 (json.dumps(d2,ensure_ascii=False), now, drv['id']))
                    conn.execute("INSERT OR REPLACE INTO car_drivers (list_id,direction,car_name,driver) VALUES (?,'arrive',?,?)",
                                 (lid, cname, drv['full']))
                    conn.execute("INSERT OR REPLACE INTO car_drivers (list_id,direction,car_name,driver) VALUES (?,'return',?,?)",
                                 (lid, cname, drv['full']))
                    assigned += 1
                    pax = 0
                    for p in pool[:]:
                        if pax >= cap-1: break
                        r = conn.execute("SELECT data FROM list_rows WHERE id=?", (p['id'],)).fetchone()
                        d2 = json.loads(r['data']); d2['רכב הלוך']=cname; d2['רכב חזור']=cname
                        conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                     (json.dumps(d2,ensure_ascii=False), now, p['id']))
                        pool.remove(p); assigned += 1; pax += 1
            conn.commit()

            ungrouped = [p['full'] for p in people if not p['arrival'] or not p['departure']]
            return jsonify({'ok':True,'assigned':assigned,'cars_used':car_idx,
                            'total_people':len(people),'groups':groups_info,'ungrouped':ungrouped})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

# ── API: Step 2 — שיבוץ הלוך ──────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/assign-step2', methods=['POST'])
@admin_required
def api_assign_step2(lid):
    from collections import defaultdict
    try:
        lst = get_list(lid)
        if not lst: return jsonify({'ok':False,'error':'רשימה לא נמצאה'})
        cap = max(1, lst['car_capacity']); car_count = max(1, lst['car_count'])
        now = datetime.now().isoformat()
        conn = get_db()
        try:
            people = _get_people(conn, lid)
            unassigned = [p for p in people if not p['car_a'] and p['arrival']]
            if not unassigned:
                return jsonify({'ok':True,'assigned':0,'message':'כולם שובצו להגעה ✅'})

            # Build pool from existing assignments
            pool_info = {}  # car_name -> {arrive_count, arrive_date, return_date}
            for i in range(1, car_count+1):
                cname = f'רכב {i}'
                pool_info[cname] = {'arrive_count':0, 'arrive_date':''}
            for p in people:
                if p['car_a'] and p['car_a'] in pool_info:
                    pool_info[p['car_a']]['arrive_count'] += 1
                    if not pool_info[p['car_a']]['arrive_date'] and p['arrival']:
                        pool_info[p['car_a']]['arrive_date'] = p['arrival']

            by_date = defaultdict(list)
            for p in unassigned: by_date[p['arrival']].append(p)

            assigned = 0
            for arr, grp in sorted(by_date.items()):
                pool = list(grp)
                # Fill existing cars with matching date
                for cname, c in pool_info.items():
                    if not pool: break
                    if c['arrive_date']==arr and c['arrive_count'] < cap:
                        while pool and c['arrive_count'] < cap:
                            p = pool.pop(0)
                            r = conn.execute("SELECT data FROM list_rows WHERE id=?", (p['id'],)).fetchone()
                            d2 = json.loads(r['data']); d2['רכב הלוך']=cname
                            conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                         (json.dumps(d2,ensure_ascii=False), now, p['id']))
                            c['arrive_count'] += 1; assigned += 1
                # Use empty cars for the rest
                for cname, c in pool_info.items():
                    if not pool: break
                    if c['arrive_count']==0 and c['arrive_date']=='':
                        drv = pool.pop(0)
                        r = conn.execute("SELECT data FROM list_rows WHERE id=?", (drv['id'],)).fetchone()
                        d2 = json.loads(r['data']); d2['רכב הלוך']=cname
                        conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                     (json.dumps(d2,ensure_ascii=False), now, drv['id']))
                        conn.execute("INSERT OR REPLACE INTO car_drivers (list_id,direction,car_name,driver) VALUES (?,'arrive',?,?)",
                                     (lid, cname, drv['full']))
                        c['arrive_count']=1; c['arrive_date']=arr; assigned+=1
                        while pool and c['arrive_count'] < cap:
                            p = pool.pop(0)
                            r = conn.execute("SELECT data FROM list_rows WHERE id=?", (p['id'],)).fetchone()
                            d2 = json.loads(r['data']); d2['רכב הלוך']=cname
                            conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                         (json.dumps(d2,ensure_ascii=False), now, p['id']))
                            c['arrive_count'] += 1; assigned += 1
            conn.commit()
            still = [p['full'] for p in _get_people(conn, lid) if not p['car_a'] and p['arrival']]
            return jsonify({'ok':True,'assigned':assigned,'still_unassigned':still})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

# ── API: Step 3 — שיבוץ חזור ──────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/assign-step3', methods=['POST'])
@admin_required
def api_assign_step3(lid):
    from collections import defaultdict
    try:
        lst = get_list(lid)
        if not lst: return jsonify({'ok':False,'error':'רשימה לא נמצאה'})
        cap = max(1, lst['car_capacity']); car_count = max(1, lst['car_count'])
        now = datetime.now().isoformat()
        conn = get_db()
        try:
            people = _get_people(conn, lid)
            unassigned = [p for p in people if not p['car_r'] and p['departure']]
            if not unassigned:
                return jsonify({'ok':True,'assigned':0,'message':'כולם שובצו לחזרה ✅'})

            # Build return pool
            pool_info = {}
            for i in range(1, car_count+1):
                cname = f'רכב {i}'
                pool_info[cname] = {'return_count':0, 'return_date':''}
            for p in people:
                if p['car_r'] and p['car_r'] in pool_info:
                    pool_info[p['car_r']]['return_count'] += 1
                    if not pool_info[p['car_r']]['return_date'] and p['departure']:
                        pool_info[p['car_r']]['return_date'] = p['departure']

            # Also check driver assignments for return direction
            drv_rows = conn.execute(
                "SELECT * FROM car_drivers WHERE list_id=? AND direction='return'", (lid,)).fetchall()
            for d in drv_rows:
                if d['car_name'] in pool_info and not pool_info[d['car_name']]['return_date']:
                    drv_p = next((p for p in people if p['full']==d['driver'] and p['departure']),None)
                    if drv_p:
                        pool_info[d['car_name']]['return_date'] = drv_p['departure']

            assigned = 0; stuck = []
            for p in unassigned:
                ret = p['departure']; placed = False
                for cname, c in pool_info.items():
                    if c['return_date']==ret and c['return_count'] < cap:
                        r = conn.execute("SELECT data FROM list_rows WHERE id=?", (p['id'],)).fetchone()
                        d2 = json.loads(r['data']); d2['רכב חזור']=cname
                        conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                     (json.dumps(d2,ensure_ascii=False), now, p['id']))
                        c['return_count'] += 1; assigned += 1; placed = True; break
                if not placed:
                    for cname, c in pool_info.items():
                        if c['return_date']=='' and c['return_count']==0:
                            r = conn.execute("SELECT data FROM list_rows WHERE id=?", (p['id'],)).fetchone()
                            d2 = json.loads(r['data']); d2['רכב חזור']=cname
                            conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                         (json.dumps(d2,ensure_ascii=False), now, p['id']))
                            c['return_date']=ret; c['return_count']=1
                            assigned += 1; placed = True; break
                if not placed:
                    stuck.append({'name':p['full'],'return_date':ret,'car_arrive':p['car_a'] or '—'})
            conn.commit()
            if stuck:
                return jsonify({'ok':False,'assigned':assigned,'stuck':stuck,
                                'stuck_count':len(stuck),'message':f'{len(stuck)} אנשים ללא רכב חזרה'})
            return jsonify({'ok':True,'assigned':assigned,'message':'כל החזרות שובצו ✅'})
        finally:
            conn.close()
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

# ── API: Reset all assignments ─────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/assign-reset', methods=['POST'])
@admin_required
def api_assign_reset(lid):
    try:
        now = datetime.now().isoformat()
        conn = get_db()
        try:
            rows = conn.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
            for row in rows:
                d = json.loads(row['data'])
                if _gf(d,'רכב חברה').lower() in ['כן','yes','true','1']: continue
                changed = False
                if d.get('רכב הלוך',''): d['רכב הלוך']=''; changed=True
                if d.get('רכב חזור',''): d['רכב חזור']=''; changed=True
                if changed:
                    conn.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                                 (json.dumps(d,ensure_ascii=False), now, row['id']))
            conn.execute("DELETE FROM car_drivers WHERE list_id=?", (lid,))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'ok':True})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

# ── API: Car View (always show full pool) ──────────────────────────────────────
@app.route('/api/lists/<int:lid>/car-view', methods=['GET'])
@auth_required
def api_car_view(lid):
    try:
        lst = get_list(lid)
        if not lst: return jsonify({'ok':False,'error':'רשימה לא נמצאה'})
        car_count    = max(1, lst['car_count'])
        car_capacity = max(1, lst['car_capacity'])
        conn = get_db()
        try:
            rows    = conn.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
            drivers = conn.execute("SELECT * FROM car_drivers WHERE list_id=?", (lid,)).fetchall()
        finally:
            conn.close()

        drv_map = {}
        for d in drivers:
            drv_map.setdefault(d['car_name'],{})[d['direction']] = d['driver']

        # Initialize ALL pool cars (empty)
        cars_arrive, cars_return = {}, {}
        for i in range(1, car_count+1):
            cname = f'רכב {i}'
            drv   = drv_map.get(cname,{}).get('arrive','') or drv_map.get(cname,{}).get('return','')
            cars_arrive[cname] = {'date':'','people':[],'driver':drv,'capacity':car_capacity,'free':car_capacity}
            cars_return[cname] = {'date':'','people':[],'driver':drv,'capacity':car_capacity,'free':car_capacity}

        unassigned_arrive, unassigned_return = [], []
        for row in rows:
            d    = json.loads(row['data'])
            if _gf(d,'רכב חברה').lower() in ['כן','yes','true','1']: continue
            full = (_gf(d,'שם')+' '+_gf(d,'שם משפחה')).strip()
            if not full: continue
            arr_date = _norm_date(_gf(d,'תאריך הגעה','הגעה'))
            ret_date = _norm_date(_gf(d,'תאריך חזרה','חזרה'))
            car_a    = _gf(d,'רכב הלוך')
            car_r    = _gf(d,'רכב חזור')

            if car_a and car_a in cars_arrive:
                c = cars_arrive[car_a]
                if full not in c['people']: c['people'].append(full)
                if arr_date and not c['date']: c['date'] = arr_date
            elif arr_date and not car_a:
                unassigned_arrive.append(full)

            if car_r and car_r in cars_return:
                c = cars_return[car_r]
                if full not in c['people']: c['people'].append(full)
                if ret_date and not c['date']: c['date'] = ret_date
            elif ret_date and not car_r:
                unassigned_return.append(full)

        # Ensure driver in both directions
        for i in range(1, car_count+1):
            cname = f'רכב {i}'
            drv = drv_map.get(cname,{}).get('arrive','') or drv_map.get(cname,{}).get('return','')
            if drv:
                if drv not in cars_arrive[cname]['people']: cars_arrive[cname]['people'].insert(0,drv)
                if drv not in cars_return[cname]['people']: cars_return[cname]['people'].insert(0,drv)

        for c in list(cars_arrive.values())+list(cars_return.values()):
            c['free'] = max(0, c['capacity']-len(c['people']))

        rec = []
        if unassigned_arrive: rec.append(f'הגעה: {len(unassigned_arrive)} ללא רכב')
        if unassigned_return: rec.append(f'חזרה: {len(unassigned_return)} ללא רכב')

        return jsonify({
            'cars_arrive':cars_arrive, 'cars_return':cars_return,
            'car_capacity':car_capacity, 'car_count':car_count,
            'unassigned_arrive':sorted(unassigned_arrive),
            'unassigned_return':sorted(unassigned_return),
            'total_arrive':sum(len(v['people']) for v in cars_arrive.values()),
            'total_return':sum(len(v['people']) for v in cars_return.values()),
            'recommendation':rec,
        })
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500


# ── API: Recommend Drivers (legacy — kept for compatibility) ──────────────────
@app.route('/api/lists/<int:lid>/recommend-drivers', methods=['GET'])
@auth_required
def api_recommend_drivers(lid):
    """
    For each arrive car, recommend the best driver:
    - The driver's return date should match as many other passengers as possible
    - So passengers can ride back with the same car
    """
    with get_db() as db:
        rows    = db.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
        drivers = db.execute("SELECT * FROM car_drivers WHERE list_id=? AND direction='arrive'", (lid,)).fetchall()

    existing_drivers = {d['car_name']: d['driver'] for d in drivers}

    # Build person info
    person_info = {}  # full_name -> {arrive_car, arrive_date, return_date}
    for row in rows:
        d = json.loads(row['data'])
        has_car = str(d.get('רכב חברה','')).lower() in ['כן','yes','true','1']
        full = (d.get('שם','') + ' ' + d.get('שם משפחה','')).strip()
        if not full or has_car: continue
        arr_car  = d.get('רכב הלוך','').strip()
        arr_date = d.get('תאריך הגעה','').strip()
        ret_date = d.get('תאריך חזרה','').strip()
        if full not in person_info and arr_car:
            person_info[full] = {'car': arr_car, 'arrive_date': arr_date, 'return_date': ret_date}

    # Group passengers by arrive car
    from collections import defaultdict
    car_passengers = defaultdict(list)  # car_name -> [person_info]
    for name, info in person_info.items():
        car_passengers[info['car']].append({'name': name, 'return_date': info['return_date']})

    recommendations = {}
    for car_name, passengers in car_passengers.items():
        if not passengers: continue
        # For each candidate driver, count how many passengers share their return date
        best_driver = None
        best_score  = -1
        best_reason = ''
        for candidate in passengers:
            cret = candidate['return_date']
            if not cret: continue
            # How many OTHER passengers return on same date?
            same_date = sum(1 for p in passengers if p['return_date'] == cret and p['name'] != candidate['name'])
            # Score = number of passengers they can bring back
            score = same_date
            if score > best_score:
                best_score  = score
                best_driver = candidate['name']
                best_reason = f'חוזר ב-{cret} עם עוד {same_date} נוסעים' if same_date > 0 else f'חוזר ב-{cret}'

        if best_driver:
            recommendations[car_name] = {
                'driver': best_driver,
                'reason': best_reason,
                'already_set': existing_drivers.get(car_name,''),
                'return_coverage': {
                    rd: [p['name'] for p in passengers if p['return_date']==rd]
                    for rd in set(p['return_date'] for p in passengers if p['return_date'])
                }
            }

    return jsonify({'recommendations': recommendations})



# ── API: Car Assignment RETURN ─────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/assign-cars-return', methods=['POST'])
@admin_required
def api_assign_cars_return(lid):
    """Phase 2: assign return trips based on driver return dates."""
    lst = get_list(lid)
    car_capacity = max(1, lst['car_capacity'])
    with get_db() as db:
        rows    = db.execute("SELECT * FROM list_rows WHERE list_id=?", (lid,)).fetchall()
        drivers = db.execute(
            "SELECT * FROM car_drivers WHERE list_id=? AND direction='arrive'", (lid,)).fetchall()

    # Map: full_name -> data
    person_data = {}
    for row in rows:
        d = json.loads(row['data'])
        full = (d.get('שם','') + ' ' + d.get('שם משפחה','')).strip()
        if full: person_data[full] = d

    # car_name -> (driver_name, driver_return_date)
    # Also inherit arrive drivers for return assignment
    car_info = {}
    for drv in drivers:
        dname = drv['driver']
        if dname and dname in person_data:
            ret = person_data[dname].get('תאריך חזרה','').strip()
            car_info[drv['car_name']] = {'driver': dname, 'return_date': ret}

    # Assign return slots: ret_date -> list of cars with drivers returning that day
    from collections import defaultdict
    cars_by_return = defaultdict(list)  # return_date -> [car_name, ...]
    for cname, info in car_info.items():
        if info['return_date']:
            cars_by_return[info['return_date']].append(cname)

    # seat counts per car (return direction)
    seat_count = defaultdict(int)
    # driver takes one seat
    for cname, info in car_info.items():
        seat_count[cname] = 1

    assignments = {}   # full_name -> car_name
    errors = []
    seen = set()

    for row in rows:
        d = json.loads(row['data'])
        has_car = str(d.get('רכב חברה','')).lower() in ['כן','yes','true','1']
        if has_car: continue
        full     = (d.get('שם','') + ' ' + d.get('שם משפחה','')).strip()
        ret_date = d.get('תאריך חזרה','').strip()
        if not full or not ret_date or full in seen: continue
        seen.add(full)

        # Skip if this person IS a driver (they already have a car)
        is_driver = any(info['driver']==full for info in car_info.values())
        if is_driver:
            # Find their car
            for cname, info in car_info.items():
                if info['driver'] == full:
                    assignments[full] = cname; break
            continue

        # Find car with driver returning same date with available seats
        found = None
        for cname in cars_by_return.get(ret_date, []):
            if seat_count[cname] < car_capacity:
                found = cname; break

        if found:
            assignments[full] = found
            seat_count[found] += 1
        else:
            available_cars = cars_by_return.get(ret_date, [])
            if available_cars:
                errors.append({'person':full,'date':ret_date,'type':'full',
                    'message':f'{full} חוזר ב-{ret_date} — כל הרכבים מאותו תאריך מלאים',
                    'suggest_extra_car': True})
            else:
                errors.append({'person':full,'date':ret_date,'type':'no_driver',
                    'message':f'{full} חוזר ב-{ret_date} — אין נהג שחוזר באותו תאריך',
                    'suggest_driver': full, 'suggest_extra_car': True})

    # Write
    updated = 0
    with get_db() as db:
        for row in rows:
            d = json.loads(row['data'])
            has_car = str(d.get('רכב חברה','')).lower() in ['כן','yes','true','1']
            full = (d.get('שם','') + ' ' + d.get('שם משפחה','')).strip()
            if has_car or full not in assignments: continue
            nv = assignments[full]
            if d.get('רכב חזור') != nv:
                d['רכב חזור'] = nv
                db.execute("UPDATE list_rows SET data=?,updated_at=? WHERE id=?",
                           (json.dumps(d,ensure_ascii=False),datetime.now().isoformat(),row['id']))
                updated += 1
        db.commit()

    return jsonify({'ok':True,'updated':updated,'errors':errors,
                    'assigned':len(assignments),'unassigned':len(errors)})


# ── API: Car View ──────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/car-driver', methods=['POST'])
@auth_required
def api_set_car_driver(lid):
    d = request.json
    with get_db() as db:
        for direction in ['arrive', 'return']:
            if d.get('driver'):
                db.execute("INSERT OR REPLACE INTO car_drivers (list_id,direction,car_name,driver) VALUES (?,?,?,?)",
                           (lid, direction, d['car_name'], d['driver']))
            else:
                db.execute("DELETE FROM car_drivers WHERE list_id=? AND direction=? AND car_name=?",
                           (lid, direction, d['car_name']))
        db.commit()
    return jsonify({'ok':True})

# ── API: Export ────────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/export', methods=['GET'])
@auth_required
def api_export(lid):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error':'pip install openpyxl'}),500
    admin = is_admin()
    with get_db() as db:
        lt   = db.execute("SELECT * FROM lists WHERE id=?", (lid,)).fetchone()
        rows = db.execute("SELECT * FROM list_rows WHERE list_id=? ORDER BY id", (lid,)).fetchall()
    cols = [c for c in LOGISTICS_COLS if admin or c['name'] not in ADMIN_ONLY_COLS]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = lt['name'][:31]; ws.sheet_view.rightToLeft=True
    hf=Font(bold=True,color='FFFFFF',size=11)
    hfill=PatternFill('solid',fgColor='1E3A5F')
    ha=Alignment(horizontal='center',vertical='center',wrap_text=True)
    thin=Side(style='thin',color='CCCCCC'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
    headers=['אישור','אושר ע"י']+[c['name'] for c in cols]
    for ci,h in enumerate(headers,1):
        cell=ws.cell(row=1,column=ci,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=ha; cell.border=border
    ws.row_dimensions[1].height=28
    fe=PatternFill('solid',fgColor='F0F4FA'); fo=PatternFill('solid',fgColor='FFFFFF'); fg=PatternFill('solid',fgColor='D4EDDA')
    for ri,row in enumerate(rows,2):
        data=json.loads(row['data']); confirmed=bool(row['confirmed_by'])
        fill=fg if confirmed else (fe if ri%2==0 else fo)
        ws.cell(row=ri,column=1,value='✓ מאושר' if confirmed else '').fill=fill; ws.cell(row=ri,column=1).border=border
        ws.cell(row=ri,column=2,value=row['confirmed_by'] or '').fill=fill; ws.cell(row=ri,column=2).border=border
        for ci,col in enumerate(cols,3):
            val=data.get(col['name'],'')
            if isinstance(val,bool): val='כן' if val else 'לא'
            cell=ws.cell(row=ri,column=ci,value=str(val) if val else '')
            cell.alignment=Alignment(horizontal='right',vertical='center'); cell.border=border; cell.fill=fill
    for ci in range(1,len(headers)+1):
        ml=max((len(str(ws.cell(row=ri,column=ci).value or '')) for ri in range(1,len(rows)+2)),default=0)
        ws.column_dimensions[get_column_letter(ci)].width=min(max(ml+4,10),35)
    ws.freeze_panes='A2'
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fn=f"{lt['name']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fn)

# ── API: Import ────────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/import', methods=['POST'])
@auth_required
def api_import(lid):
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error':'pip install openpyxl'}),500
    if 'file' not in request.files: return jsonify({'error':'לא נשלח קובץ'}),400
    f=request.files['file']
    if not f.filename.lower().endswith(('.xlsx','.xls')): return jsonify({'error':'xlsx בלבד'}),400
    try:
        wb=openpyxl.load_workbook(io.BytesIO(f.read()),read_only=True,data_only=True)
        ws=wb.active; rows_iter=list(ws.iter_rows(values_only=True))
    except Exception as ex: return jsonify({'error':str(ex)}),400
    if len(rows_iter)<2: return jsonify({'error':'קובץ ריק'}),400
    excel_headers=[normalize_header(h) for h in rows_iter[0]]
    norm_cols={normalize_header(c['name']):c['name'] for c in LOGISTICS_COLS}
    col_map={ei:norm_cols[hdr] for ei,hdr in enumerate(excel_headers) if hdr in norm_cols}
    if not col_map: return jsonify({'error':f'אין עמודות תואמות. בקובץ: {excel_headers[:5]}'}),400
    imported=0
    with get_db() as db:
        for row_vals in rows_iter[1:]:
            if not any(v is not None and str(v).strip() for v in row_vals): continue
            data={c['name']:'' for c in LOGISTICS_COLS}
            for ei,col_name in col_map.items():
                if col_name in ADMIN_ONLY_COLS and not is_admin(): continue
                data[col_name]=cell_to_str(row_vals[ei] if ei<len(row_vals) else None)
            db.execute("INSERT INTO list_rows (list_id,data) VALUES (?,?)",
                       (lid,json.dumps(data,ensure_ascii=False)))
            imported+=1
        db.commit()
    # Auto-detect config
    with get_db() as db:
        all_rows=db.execute("SELECT data FROM list_rows WHERE list_id=?", (lid,)).fetchall()
        lst=db.execute("SELECT * FROM lists WHERE id=?", (lid,)).fetchone()
    old_teams=set(json.loads(lst['teams']))
    teams_found=set(); dates_a=[]; dates_r=[]; no_car=0
    for r in all_rows:
        d=json.loads(r['data'])
        t=d.get('צוות','').strip()
        if t: teams_found.add(t)
        arr=d.get('תאריך הגעה','').strip(); dep=d.get('תאריך חזרה','').strip()
        if arr and len(arr)==10: dates_a.append(arr)
        if dep and len(dep)==10 and dep!='2027-01-01': dates_r.append(dep)
        if str(d.get('רכב חברה','')).lower() not in ['כן','yes','true','1']: no_car+=1
    new_teams=sorted(old_teams|teams_found)
    t_start=min(dates_a) if dates_a else (lst['trial_start'] or '')
    t_end  =max(dates_r) if dates_r else (lst['trial_end']   or '')
    sugg_cars=max(1,math.ceil(no_car/3)) if no_car else (lst['car_count'] or 4)
    with get_db() as db:
        db.execute("UPDATE lists SET teams=?,car_count=?,trial_start=?,trial_end=? WHERE id=?",
                   (json.dumps(new_teams,ensure_ascii=False),sugg_cars,t_start or None,t_end or None,lid))
        db.commit()
    suggestions={'teams':new_teams,'new_teams':sorted(teams_found-old_teams),
                 'trial_start':t_start,'trial_end':t_end,'car_count':sugg_cars,'no_car_count':no_car,'applied':True}
    return jsonify({'ok':True,'imported':imported,'matched_cols':list(col_map.values()),'suggestions':suggestions})

@app.route('/api/lists/<int:lid>/excel-preview', methods=['POST'])
@auth_required
def api_excel_preview(lid):
    try:
        import openpyxl
    except ImportError:
        return jsonify({'error':'pip install openpyxl'}),500
    if 'file' not in request.files: return jsonify({'error':'לא נשלח קובץ'}),400
    f=request.files['file']
    try:
        wb=openpyxl.load_workbook(io.BytesIO(f.read()),read_only=True,data_only=True)
        sheets=[]
        for sname in wb.sheetnames:
            ws=wb[sname]; rows=list(ws.iter_rows(values_only=True,max_row=3))
            headers=[normalize_header(h) for h in (rows[0] if rows else []) if h]
            sample=[cell_to_str(v) for v in rows[1][:len(headers)]] if len(rows)>1 else []
            total=sum(1 for _ in ws.iter_rows(min_row=2,values_only=True) if any(v is not None for v in _))
            sheets.append({'name':sname,'headers':headers,'sample':sample,'total':total})
        wb.close()
    except Exception as ex: return jsonify({'error':str(ex)}),400
    return jsonify({'sheets':sheets})

# ── API: Changelog ─────────────────────────────────────────────────────────────
@app.route('/api/lists/<int:lid>/changelog', methods=['GET'])
@auth_required
def api_changelog(lid):
    with get_db() as db:
        logs=db.execute("SELECT * FROM changes_log WHERE list_id=? ORDER BY created_at DESC LIMIT 200",(lid,)).fetchall()
    return jsonify([dict(l) for l in logs])

@app.route('/api/lists/<int:lid>/changelog-clear', methods=['POST'])
@admin_required
def api_changelog_clear(lid):
    try:
        conn = get_db()
        try:
            conn.execute("DELETE FROM changes_log WHERE list_id=?", (lid,))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'ok':True})
    except Exception as e:
        return jsonify({'ok':False,'error':str(e)}), 500

# ── API: Users (admin only) ────────────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@admin_required
def api_get_users():
    with get_db() as db:
        users = db.execute("SELECT id,name,username,role,approved,created_at FROM users ORDER BY id").fetchall()
    return jsonify([dict(u) for u in users])

@app.route('/api/users', methods=['POST'])
@admin_required
def api_create_user():
    d = request.json
    name     = d.get('name','').strip()
    username = d.get('username','').strip()
    password = d.get('password','').strip()
    if not name or not password:
        return jsonify({'error':'שם וסיסמה הם שדות חובה'}), 400
    if not username:
        username = name
    try:
        with get_db() as db:
            cur = db.execute("INSERT INTO users (name,username,password,role,approved) VALUES (?,?,?,?,?)",
                             (name, username, password, 'user', 1))
            db.commit()
        return jsonify({'ok':True,'id':cur.lastrowid})
    except Exception as e:
        return jsonify({'error':str(e)}), 400

@app.route('/api/users/<int:uid>', methods=['PUT'])
@admin_required
def api_update_user(uid):
    d = request.json
    with get_db() as db:
        if d.get('password'):
            db.execute("UPDATE users SET name=?,password=? WHERE id=?",
                       (d['name'], d['password'], uid))
        else:
            db.execute("UPDATE users SET name=? WHERE id=?", (d['name'], uid))
        db.commit()
    return jsonify({'ok':True})

@app.route('/api/users/<int:uid>', methods=['DELETE'])
@admin_required
def api_delete_user(uid):
    with get_db() as db:
        db.execute("DELETE FROM users WHERE id=?", (uid,))
        db.commit()
    return jsonify({'ok':True})



# ── API: Approve / Reject User ─────────────────────────────────────────────────
@app.route('/api/users/<int:uid>/approve', methods=['POST'])
@admin_required
def api_approve_user(uid):
    with get_db() as db:
        db.execute("UPDATE users SET approved=1 WHERE id=?", (uid,))
        db.commit()
    return jsonify({'ok': True})

@app.route('/api/users/<int:uid>/reject', methods=['POST'])
@admin_required
def api_reject_user(uid):
    with get_db() as db:
        db.execute("DELETE FROM users WHERE id=?", (uid,))
        db.commit()
    return jsonify({'ok': True})


# ── API: Settings ──────────────────────────────────────────────────────────────
@app.route('/api/settings', methods=['POST'])
@admin_required
def api_settings():
    d=request.json
    with get_db() as db:
        for k in ['password','admin_password']:
            if d.get(k): db.execute("INSERT OR REPLACE INTO app_settings VALUES (?,?)",(k,d[k]))
        db.commit()
    return jsonify({'ok':True})

if __name__=='__main__':
    import socket
    try: local_ip=socket.gethostbyname(socket.gethostname())
    except: local_ip='127.0.0.1'
    print(f"\n{'='*55}\n  🗂️  ניהול לוגיסטיקה v13\n  http://localhost:{PORT}\n  http://{local_ip}:{PORT}\n  סיסמה: logistics2026 | אדמין: admin2026\n{'='*55}\n")
    app.run(host='0.0.0.0',port=PORT,debug=False)
